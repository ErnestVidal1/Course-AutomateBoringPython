'''
Program: Number Guessing Game

Author: Ernest Vidal

Description: This program is a number-guessing game that serves as a practical application of various programming concepts. It incorporates examples from different sections of the Udemy course "Automate the Boring Stuff with Python Programming" by instructor Al Sweigart. The game aims to offer an interactive way to apply and understand Python programming constructs ranging from basic syntax to more advanced features like file operations and GUI automation.

##########################
# TABLE OF CONTENTS
##########################
Section 1: Python Basic (Subsections 1, 2, 3)
Section 2: Flow Control (Subsections 4, 5, 6, 7)
Section 3: Functions (Subsections 8, 9, 10)
Section 4: Handling Errors with try/except (Subsection 11)
Section 5: Writing a Complete Program: Guess the Number (Subsection 12)
Section 6: Lists (Subsections 13, 14, 15, 16)
Section 7: Dictionaries (Subsections 17, 18)
Section 8: More About Strings (Subsections 19, 20, 21)
Section 9: Running Programs from the Command line (Subsection 22)
Section 10: Regular Expressions (Subsections 23, 24, 25, 26, 27, 28, 29)
Section 11: Files (Subsections 30, 31, 32, 33, 34)
Section 12: Debugging (Subsections 35, 36, 37)
Section 13: Web Scraping (Subsections 38, 39, 40, 41)
Section 14: Excel, Word, and PDF Documents (Subsections 42, 43, 44, 45)
Section 15: Email (Subsections 46, 47)
Section 16: GUI automation (Subsections 48, 49, 50, 51)

##########################
# SUMMARY
##########################
1. Overview:
   - This program is a number guessing game. The player has 7 turns to guess a randomly generated number within a single game series.
  
2. Architecture:
    - Main Loop: Manages the overall flow, prompting the player to start or restart the game.
    - Game Series Loop ("for" loop): Manages a single game series, which consists of 7 turns.
    - Validation Loop ("while" loop): Validates that the user's input is a valid number for a turn (or guess).

3. Components:
    - 'number_frequency' Dictionary: Tracks the frequency of numbers guessed across multiple game series.
    - 'guesses' List: Stores the player's guesses within a single game series.
  
4. Workflow:
    - The Main Loop initializes the program and asks the player if they wish to play.
    - Upon confirmation, a Game Series Loop starts, limited to 7 turns.
    - During each turn, the Validation Loop ensures the input is a valid and unique number.
    - The 'guesses' list keeps track of guesses for the current game series.
    - The 'number_frequency' dictionary accumulates data on how often each number is guessed across game series.

5. Course References and Comments:
    - Throughout the program, references are made to various sections of the course. These references highlight different Python topics (e.g., For Loops, While Loops, List Methods, etc.). Alongside these course references, there are explanatory comments that clarify the program's functionality step-by-step.

'''


##########################
# IMPORTS
##########################


import random
import time
from pprint import pprint
import os
import subprocess
import re
import shutil
import webbrowser
import requests
import logging
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import PyPDF2
from docx import Document
import smtplib


'''
COURSE: Subsection 36 (=12.2). Logging
'''
# Configure logging settings for debugging and tracking. Logs are saved in "my_program.log" and indicate the program's start.


logging.basicConfig(filename='my_program.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logging.info('Program started.')


##########################
# INITIALIZE GAME SETTINGS 
##########################


logging.info("Initializing game settings...")


# ~~~~~~~~~~~~~~~~~~~~~~~~
# DEBUGGING SETUP
# ~~~~~~~~~~~~~~~~~~~~~~~~


'''
COURSE: Subsection 37 (=12.3). Using the Debugger
''' 
# For debugging with pdb, uncomment the following line


# import pdb; pdb.set_trace()


logging.debug("Debugging setup initialized.")
    
    
##########################
# GAME: GUESS THE NUMBER 
##########################


logging.info("Guess the Number game started.")


'''
COURSE: Subsection 8 (=3.1). Def, return and None
'''
# Define the main function for playing the guess the number game.


def play_guess_the_number(number_frequency): 


    # ~~~~~~~~~~~~~~~~~~~~~~~~
    # GAME INITIALIZATION
    # ~~~~~~~~~~~~~~~~~~~~~~~~


    '''
    COURSE: Subsection 2 (=1.2). Basic Terminology and Using IDLE
    ''' 
    # Initialize the secret number with a random value between 1 and 20.
    # Perform mathematical transformations on the secret number for demonstration purposes related to the course subsection. These operations are not necessary for the game logic.
    
    secretNumber = random.randint(1, 20)     
    secretNumber = (secretNumber - 1) * ((7 + 1) / (3 - 1))
    secretNumber = int((secretNumber % 20) + 1)
    
    
    '''
    COURSE: Subsection 13 (=6.1). The List Data Type
    '''
    # Initialize the 'guesses' list to store the player's guesses for the current game series.
    # This list will be emptied at the beginning of each new game series.
    # Its purpose is to track the numbers guessed by the player during a single series of the game,
    # to ensure that the same number is not guessed more than once in that series.

    
    guesses = []
    
    
    '''
    COURSE: Subsection 3 (=1.3). Writing Our First Program
    ''' 
    # Display initial game message.
    
    
    print('I am thinking of a number between 1 and 20.')

    
    '''
    COURSE: Subsection 2 (=1.2). Basic Terminology and Using IDLE
    ''' 
    # Initialize start_time to capture the elapsed time for player's guesses. Used later to display player performance.


    start_time = time.time()  
    
    logging.info("Game variables and configurations initialized.")
    
    
    '''
    COURSE: Subsection 7 (=2.4). For Loops
    ''' 
    # Initialize a loop to control the number of turns (guess attempts) in each game series, with a maximum of 7 turns per game series.


    # ~~~~~~~~~~~~~~~~~~~~~~~~
    # GAME SERIES LOOP
    # ~~~~~~~~~~~~~~~~~~~~~~~~
    
    
    logging.debug("Entering game series loop.")
    
    for guessesTaken in range(1, 7):
    
    
        '''
        COURSE: Subsection 6 (=2.3). While Loops
        '''

        # Start an inner loop for input validation. This loop will continue asking for a valid integer input 
        # until one is provided and has not been guessed before in the current game series. The loop will exit (break) under these conditions.
        
        
        # ~~~~~~~~~~~~~~~~~~~~~~~~
        # VALIDATION LOOP
        # ~~~~~~~~~~~~~~~~~~~~~~~~
    
    
        logging.debug("Entering validation loop.")


        while True:
        
        
            '''
            COURSE: Subsection 11 (=4.1). Try and Except Statements
            '''
            # Use try-except to catch invalid inputs that are not numbers.
        
        
            try:
            
            
                '''
                COURSE: Subsection 3 (=1.3). Writing Our First Program
                ''' 
                # Prompt the user to make a guess. This is the basic input/output feature of the program.


                guess = input('Take a guess: ')
                
                
                '''
                COURSE: Subsection 20 (=8.2). String Methods (strip)
                '''
                # Remove any leading and trailing white spaces from the user input for cleaner data handling.
                
                
                guess = guess.strip()
                
                
                # Try to convert input to integer. If it fails, it triggers 'except ValueError'.
                

                guess = int(guess)


                # If successful, check if number was already guessed. If so, ask for a new guess.

                
                if guess in guesses:
                    print("You've already guessed that number. Try again.")
                    
                    
                    # The 'continue' statement applies to the closest loop, which is the inner 'while' loop.
                    
                    
                    continue
                else:
                
                
                    '''
                    COURSE: Subsection 15 (=6.3). List Methods (append)
                    ''' 
                    # If the guess is unique (not in the 'guesses' list), append it to the list for future reference.


                    guesses.append(guess)                    
                    
                    
                    '''
                    COURSE: Subsection 17 (=7.1). The Dictionary Data Type
                    ''' 
                    # The 'number_frequency' dictionary was initialized at the start of the main program with 'number_frequency = {}'.
                    # It tracks how often each number is guessed across multiple game series (each series managed by a "for" loop and consisting of seven turns) as long as the program is running.
                    # E.g., initial number_frequency, note the number 3 = {3: 1, 7: 1, 18: 4}.
                    # The 'get()' method retrieves the current frequency of the guessed number (guess), defaulting to 0 if it's a new number: number_frequency.get(guess, 0).
                    # Then, the frequency count for that guessed number is incremented by 1: + 1.
                    # Finally, the new frequency is assigned back to the corresponding dictionary key: number_frequency[guess] = ...
                    # E.g., after updating, note the number 3 = {3: 2, 7: 1, 18: 4}.
                    # When you start a new game series without restarting the program, upon the complete finalization of a "for" loop, the dictionary retains its previous state.

                    
                    number_frequency[guess] = number_frequency.get(guess, 0) + 1
                    
                    
                    # Exit the Validation Loop ("while" loop) since a valid and unique number has been guessed.
                    # This allows the program to proceed to evaluating the guess against the target number.


                    break
                    
                    
            # Display an error message in relation to the 'try' block above when the input is not a valid number.
                    
                    
            except ValueError:     
            
                logging.error("ValueError encountered: Invalid input, not a number.")
                
                print('Please enter a number.')
                
                
                
                
# !!!!!!!!!!!!!!!!!!!!!!!!    
# !!!!!!!!!!!!!!!!!!!!!!!!
# !!! TODO: Starting from this line, continue adding section-related comments for the course. !!!
# !!!!!!!!!!!!!!!!!!!!!!!!
# !!!!!!!!!!!!!!!!!!!!!!!!


        # ~~~~~~~~~~~~~~~~~~~~~~~~
        # VALIDATION LOOP
        # ~~~~~~~~~~~~~~~~~~~~~~~~
    
    
        logging.debug("Entering validation loop.")


        while True:
        
        
        
        
        
        
        
        
        if guess < secretNumber:
            print('Your guess is too low.')
        elif guess > secretNumber:
            print('Your guess is too high.')
        else:
            # Section 3.2: Keyword Arguments and print()
            # Section 3.3: Local and Global Scope
            end_time = time.time()
            elapsed_time = end_time - start_time
            # Section 10.1. Regular Expressions Basics
            # Section 10.5. Regex Dot-Star and the Caret/Dollar Characters
            while True:
                name = input("What's your name? ")
                if re.fullmatch('^[A-Za-z]+$', name):
                    break
                else:
                    print("Invalid name. Please enter a name that contains only letters.")
            # Section 10.2. Regex Groups and the Pipe Character
            # Section 10.6.Regex sub() Method and Verbose Mode
            # Section 12.1. The raise and assert Statements
            # Define la excepción
            class InvalidEmailError(Exception):
                pass
            email_regex = re.compile(r'''
                ([a-zA-Z0-9._%+-]+)  # Nombre de usuario
                (@)                  # Símbolo arroba
                ([a-zA-Z0-9.-]+)     # Dominio
                (\.[a-zA-Z]{2,})     # TLD
            ''', re.VERBOSE)

            while True:
                try:
                    email = input("What's your email? (or type 'N/A' if you don't want to provide) ")
                    if email_regex.fullmatch(email):
                        break
                    elif email == 'N/A':
                        break
                    else:
                        # Section 12.1. The raise and assert Statements
                        raise InvalidEmailError("Invalid email. Please enter a valid email address or 'N/A'.")
                except InvalidEmailError as e:
                    print(e)
            # Reemplazamos todo excepto el dominio con asteriscos.
            masked_email = email_regex.sub(r'****\2\3\4', email)
            print(f"For security reasons, only this part of your email is stored: {masked_email}")
            # Section 10.3. Repetition in Regex Patterns and Greedy/Nongreedy Matching
            # This regex pattern allows phone numbers with optional country codes, spaces, or hyphens as separators.
            # It also allows for multiple numbers separated by commas. Examples: "123", "123-123-4444", "+1 123", "123, +4 5678, 90".
            phone_regex = re.compile(
                r'(?:\+\d{1,4}\s?)?(?:\d{1,4}\s?)*(?:,\s*(?:\+\d{1,4}\s?)?(?:\d{1,4}\s?)*)*|(N/A)')  # Valid phone numbers or 'N/A'
            while True:
                phone = input(
                    "What's your phone number(s)? You can enter multiple numbers separated by commas (or type 'N/A' if you don't want to provide) ")
                if phone_regex.fullmatch(phone):
                    break
                else:
                    print("Invalid phone number format. Optional country codes can be included.")
            # Section 8.3. String Formatting
            print("Good job, %s! You guessed my number %d in %.2f seconds." % (name, guessesTaken, elapsed_time))
            # Section 7.2: Data structures
            player_record = {'name': name, 'time': elapsed_time}  # Ahora 'name' está definido
            # Section 7.2: Data structures
            fastest_times_list.append(player_record)  # Añade el registro a la lista
            print('Fastest times in correctly guessing the number:')
            pprint(fastest_times_list)
            # Debug: Imprimir el valor de elapsed_time
            print(f"Debug: elapsed_time = {elapsed_time}")
            # Section 10.4. Regex Character Classes and the findall() Method
            # Convertir los tiempos flotantes en fastest_times_list a cadenas de texto
            time_str_list = [str(record['time']) for record in fastest_times_list]
            time_str = ', '.join(time_str_list)
            # Search for fast and slow times using regular expressions
            fast_times = re.findall(r'\b([0-4])\b', time_str)
            slow_times = re.findall(r'\b([5-9]|[1-5][0-9]|60)\b', time_str)
            print(f"You had {len(fast_times)} fast guesses (less than 5 seconds) and {len(slow_times)} slow guesses (5 seconds or greater) based on your elapsed time.")
            # Section 1.3: String Concatenation and Replication
            print((name.upper() + " WELL DONE! ") * 5)
            # Section 9. Running Programs from the Command line
            current_folder = os.path.dirname(os.path.abspath(__file__))
            bat_path = os.path.join(current_folder, 'congrats.bat')
            # Section 11.1. Filenames and Absolute/Relative File Paths
            # Verificar si el archivo 'congrats.bat' existe
            if not os.path.exists(bat_path):
                # Si no existe, crearlo
                # Section 11.2. Reading and Writing Plaintext files
                with open(bat_path, 'w') as f:
                    f.write("@echo off\n")
                    f.write('set "name=%~1"\n')
                    f.write("echo Good job, %name%!\n")
            # Ejecutar el archivo .bat
            subprocess.run([bat_path, name])
            # "Section 11.3. Copying and Moving Files and Folders
            # Crear una copia del archivo .bat
            copy_path = os.path.join(current_folder, 'congrats_copy.bat')
            shutil.copy(bat_path, copy_path)

            # Ejecutar la copia
            subprocess.run([copy_path, name])
            # Pequeña pausa para asegurarse de que el archivo se ha liberado
            time.sleep(2)
            # Section 11.4. Deleting files
            # Eliminar la copia
            # Section 11.5. Walking a Directory Tree
            for dirpath, dirnames, filenames in os.walk(os.path.dirname(copy_path)):
                if os.path.basename(copy_path) in filenames:
                    print(f"Found the file: {copy_path}")
                    logging.info('About to try unlinking the file.')
                    # Añadir una demora antes de intentar eliminar el archivo
                    time.sleep(1)
                    try:
                        os.unlink(copy_path)
                    except PermissionError as e:
                        logging.error(f"Error al intentar eliminar el archivo: {e}")
                    logging.info('Past the try/except block.')
                    break  # No need to continue walking through the directory tree

            # Section 14.4. Reading and Editing Word documents
            # Copia el archivo de la plantilla
            os.getcwd()
            while True:
                try:
                    # Intenta copiar el archivo
                    shutil.copy('template-congratulations.docx', 'congratulations.docx')
                    break  # Si la copia es exitosa, rompe el loop
                except PermissionError:
                    # Muestra un mensaje de error amigable
                    print(
                        "Error de permiso. Asegúrate de que el archivo 'congratulations.docx' no esté abierto en otra aplicación.")
                    # Pregunta al usuario si quiere intentarlo de nuevo
                    choice = input("¿Quieres intentarlo de nuevo? (s/n): ")
                    if choice.lower() != 's':
                        print("Operación cancelada por el usuario.")
                        break  # Si el usuario elige no intentarlo de nuevo, rompe el loop
            # Abre el nuevo archivo para edición
            doc = Document('congratulations.docx')
            # Reemplaza el marcador de posición con el nombre del jugador
            for paragraph in doc.paragraphs:
                if 'NAME' in paragraph.text:
                    for run in paragraph.runs:
                        run.text = run.text.replace('NAME', name)
            # Guarda el documento
            doc.save('congratulations.docx')
            # Abre el archivo Word una vez que está modificado
            os.system(f'open congratulations.docx' if os.name == 'posix' else f'start congratulations.docx')

    print(f'Sorry. The number I was thinking of was {secretNumber}.')
    print(f'Your guesses were: {", ".join(map(str, guesses))}')

# Section 1.5: Program
# 5. Section 5: Writing a Complete Program. Guess the Number
if __name__ == "__main__":
    # Section 8.1 . Advanced String Syntax
    print("""Welcome to Guess the Number game!
    Rule 1: You will have 6 attempts to guess the correct number.
    Rule 2: The number will be between 1 and 20.
    Rule 3: No repeated guesses are allowed.
    """)
    game_count = 0
    play_again = True
    # Section 7: Dictionaries
    # Section 7.1: The Dictionary Data Type
    number_frequency = {} # New dictionary to store frequency of guessed numbers
    # Section 7.2: Data structures.
    fastest_times_list = [] # List to store fastest times.

    # Section 2.4: Elements of Flow Control
    
    
    # ~~~~~~~~~~~~~~~~~~~~~~~~
    # MAIN GAME LOOP
    # ~~~~~~~~~~~~~~~~~~~~~~~~
    
    
    logging.debug("Entering main game loop.")

    
    while play_again:
        play_guess_the_number(number_frequency)  # Modified function call to pass number_frequency
        # Section 6.1: The Augmented Assignment Operators
        game_count += 1
        print(f'You have played {game_count} games.')
        # Section 7: Dictionaries
        # Section 7.1: The Dictionary Data Type
        # Section 14.1. Reading Excel Spreadsheets
        # Section 14.2. Editing Excel Spreadsheets
        # Función para guardar las frecuencias en Excel
        def save_to_excel(data):
            wb = Workbook()
            ws = wb.active
            ws.append(["Número", "Frecuencia"])
            for k, v in data.items():
                ws.append([k, v])
            wb.save("frequencies.xlsx")
        # Función para leer las frecuencias desde Excel y mostrarlas
        def read_from_excel():
            wb = load_workbook(filename="frequencies.xlsx")
            ws = wb.active
            data = {}
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=2):
                number = row[0].value
                frequency = row[1].value
                data[number] = frequency
            return data

        # Section 14.3. Reading and Editing PDFs
        def save_to_pdf():
            excel_path = 'frequencies.xlsx'
            pdf_path = 'frequencies.pdf'
            convert(excel_path, pdf_path)

        choice = input("¿Quieres ver las frecuencias en la consola o en Excel? (consola/excel): ").lower()
        if choice == "consola":
            print('Frequencies of numbers you have attempted to guess so far:')
            pprint(number_frequency)  # Utiliza pprint para mostrar las frecuencias
        elif choice == "excel":
            save_to_excel(number_frequency)  # Guardar datos en Excel
            print("Las frecuencias se han guardado en 'frequencies.xlsx'")
            os.system('open frequencies.xlsx' if os.name == 'posix' else 'start excel frequencies.xlsx')  # Abrir Excel
        else:
            print("Opción no válida.")

        while True:
            choice = input('Do you want to play again? (yes/no): ')
            # Section 8.2. String Methods (lower, startswith)
            choice = choice.lower()
            if choice.startswith('y'):
                choice = 'yes'
            elif choice.startswith('n'):
                choice = 'no'
            if choice not in ['yes', 'no']:
                print('Please answer with "yes" or "no".')
            else:
                # Section 6.1: The Augmented Assignment Operators
                play_again = choice == 'yes'
                if not play_again:
                    number_frequency = {}  # Restablece a cero el diccionario.
                    fastest_times_list = []  # Reset the list to empty.
                break

from selenium import webdriver
from bs4 import BeautifulSoup
import requests
import webbrowser

# Importa las bibliotecas necesarias para todas las secciones
import webbrowser
import requests
from bs4 import BeautifulSoup

# Descomentar la siguiente línea si planeas usar Selenium
# from selenium import webdriver

# Sección 13.1: El módulo webbrowser
while True:


    more_info = input("Would you like to know more about guessing games? (yes/no): ").lower()
    if more_info == 'yes':
        info_format = input("How would you like to receive the information? (web/pdf): ").lower()

        if info_format == 'web':
            # Abre el artículo en el navegador
            print("Opening Wikipedia article...")
            webbrowser.open('https://en.wikipedia.org/wiki/Guessing_game')
            # Sección 13.2: Descarga de la web con el módulo Requests
            read_text = input(
                "Would you like to read the initial part of the Wikipedia article on guessing games right here in the console? (yes/no): ").lower()
            if read_text == 'yes':
                print("Fetching content...")
                url = "https://en.wikipedia.org/wiki/Guessing_game"
                res = requests.get(url)
                res.raise_for_status()

                # Sección 13.3: Análisis de HTML con el módulo Beautiful Soup
                soup = BeautifulSoup(res.text, 'html.parser')
                first_paragraph_bs4 = soup.select('p')[0].getText()  # Selecciona el primer párrafo

                # Sección 13.4: Control del navegador con el módulo Selenium (comentado)
                '''
                # To utilize the Selenium functionality, make sure you have uncommented the Selenium import at the beginning of this code file.
                # Uncomment the line to import Selenium, which is required for the code below to run.
                # Here, make sure the version of Chromedriver matches the version of Chrome installed on your system.
                # You can download the matching version of Chromedriver from https://sites.google.com/a/chromium.org/chromedriver/
                # Once downloaded, place the executable in an accessible location and update the path below.
                driver_path = "tu_ruta_al_archivo_chromedriver"
                driver = webdriver.Chrome(executable_path=driver_path)
                driver.get(url)
                second_paragraph_selenium = driver.find_elements_by_css_selector('p')[1].text  # Selecciona el segundo párrafo
                driver.quit()
                '''

                print(f"Here is the first paragraph from the article:\n{first_paragraph_bs4}")
                # Descomentar la siguiente línea si activas la sección de Selenium
                # print(f"Here is the second paragraph from the article:\n{second_paragraph_selenium}")
            break

        elif info_format == 'pdf':
            # Section 14.3. Reading and Editing PDFs
            # Código para abrir el PDF aquí

            def get_pdf_page_count(pdf_path):
                with open(pdf_path, 'rb') as pdf_file:
                    pdf_reader = PyPDF2.PdfReader(pdf_file)
                    return len(pdf_reader.pages)


            pdf_path = "guide.pdf"
            page_count = get_pdf_page_count(pdf_path)
            print(f"Opening the PDF guide... It has {page_count} pages.")
            os.system(f'open {pdf_path}' if os.name == 'posix' else f'start {pdf_path}')
        else:
            print("Invalid option.")

    elif more_info == 'no':
    # Section 15.1. Sending emails
    # Section 15.2. Checking your Email inbox (no aplicado)
        def send_email(subject, body, to_email):
            from_email = "your_email@example.com"
            message = f"Subject: {subject}\n\n{body}"
            smtpObj = smtplib.SMTP('smtp.example.com', 587)
            smtpObj.ehlo()
            smtpObj.sendmail(from_email, to_email, message)
            smtpObj.quit()
        if 'email' in locals() and email != 'N/A':  # Comprobar que el usuario haya proporcionado un email
            subject = "Thanks for Playing!"
            body = f"Hi {name},\n\nWe hope you had fun playing our game. Here are your number frequencies:\n{number_frequency}"
            send_email(subject, body, email)
        print("Thank you for playing!")
        break
    else:
        print("Please enter 'yes' or 'no'.")


# Section 16. GUI automation (no aplicado).
# Section 12.2. Logging
# Agrega un registro para indicar que el programa ha finalizado
logging.info('Programa finalizado.')

    
##########################
# GAME OVER: REVEAL RESULT 
##########################


logging.debug("Game Over. Revealing results.")
